import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

class SS_manager (): 
    """Manage local spread sheets
    """
    
    def __init__(self, file_name): 
        
        self.file_name = file_name
        self.wb = openpyxl.load_workbook(self.file_name)
        self.current_sheet = None
    
    def get_sheets (self): 
        """ Return the list of sheets in the current document
        """
        
        return self.wb.sheetnames
    
        
    def clean_workbook (self): 
        """ Delete all sheets in current workbook
        """
        
        for sheet in self.wb.sheetnames: 
            sheet_obj = self.wb[sheet]
            self.wb.remove(sheet_obj)
            
    def create_get_sheet (self, sheet_name):
        """ Create a new sheet with specifici name, and set it as
        current sheet in class 
        """
        
        self.wb.create_sheet(sheet_name)
        self.set_sheet(sheet_name)
    
    def set_sheet (self, sheet_name): 
        
        self.current_sheet = self.wb[sheet_name]
    
    def save (self):
        """Save current workbook
        """
        
        self.wb.save(self.file_name)

    def write_cell (self, value="", row=1, column=1): 
        """ Write data in specific cell 
        """
        
        self.current_sheet.cell (row, column).value = value
    
    def write_data (self, data=[], start_row=1, start_column=1): 
        """ Write data list starting in specific cell
        """
                
        current_row = start_row
        current_column = start_column
        
        for row in data: 
                        
            for cell_value in row: 
                                
                cell_obj = self.current_sheet.cell (current_row, current_column)
                cell_obj.value = cell_value
                
                current_column += 1
            
            current_column = start_column
            current_row += 1
    
    def auto_width (self): 
        """ Set corect width to each coumn in the current sheet
        """
    
        for col in self.current_sheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.current_sheet.column_dimensions[column].width = adjusted_width
        
    def format_range (self, start_cell=(1,1), end_cell=(1,1), italic=False, 
                      bold=False, font_size=8): 
        
        # Create font style
        formated_font = Font(size=font_size, italic=italic, bold=bold)
        
        # Apply style
        current_row = start_cell[0]
        current_column = start_cell[1]
        
        for row in range(start_cell[0], end_cell[0] + 1): 
                        
            for cell_value in range(start_cell[1], end_cell[1] + 1): 
                                
                cell_obj = self.current_sheet.cell (current_row, current_column)
                cell_obj.font = formated_font
                
                current_column += 1
            
            current_column = 1
            current_row += 1
        
    def get_data (self): 
        """ Get all data from the current page """

        rows = self.current_sheet.max_row
        columns = self.current_sheet.max_column

        data = []
        for row in range(1, rows + 1):

            row_data = []
            for column in range(1, columns + 1):
                cell_data = self.current_sheet.cell (row, column).value
                row_data.append (cell_data)

            data.append (row_data)

        return data